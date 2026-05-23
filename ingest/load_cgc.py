"""
Bulk-load every CGC GSW weekly file in data/raw/cgc/ into cgc_weekly.

EXPECTED LAYOUT:
    data/raw/cgc/
        2016-2017/
            gsw-shg-1-en.xls
            gsw-shg-2-en.xls
            ...
            gsw-shg-52-en.xls
        2017-2018/
            ...
        ...
        2025-2026/
            ...

The folder name (e.g. "2016-2017") is the crop year. Files inside that
folder MUST belong to that crop year — we use the folder name as the
authoritative source. We still parse the title text inside the file to
extract the week-ending date, but if the file's date contradicts the folder
we trust the folder and log a warning.

Handles both .xls and .xlsx automatically.
Handles combined-week files (Christmas weeks etc) — splits into 2 rows.
Re-running is safe — rows are upserted by (week_ending, geography, crop).

Failures are logged to data/raw/cgc_load_failures.log instead of crashing
the whole batch.

Run:
    python3 ingest/load_cgc.py
"""

import re
import sqlite3
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook   # .xlsx
import xlrd                          # .xls

CGC_DIR = Path(__file__).parent.parent / "data" / "raw" / "cgc"
DB_PATH = Path(__file__).parent.parent / "data" / "db" / "canola.sqlite"
FAIL_LOG = Path(__file__).parent.parent / "data" / "raw" / "cgc_load_failures.log"

SECTIONS = [
    ("Primary", "Producer Deliveries to Primary Elevators -",       "producer_deliveries_weekly_kt"),
    ("Primary", "Primary Elevator Shipments -",                     "primary_shipments_weekly_kt"),
    ("Primary", "Crop Year to Date Producer Deliveries",            "producer_deliveries_cytd_kt"),
    ("Primary", "Crop Year to Date Primary Elevator Shipments",     "primary_shipments_cytd_kt"),
    ("Primary", "Stocks at Primary Elevators",                      "primary_elevator_stocks_kt"),
    ("Primary", "Condo Storage at Primary Elevators",               "condo_storage_kt"),
    ("Process", "Producer Deliveries to Process",                   "process_deliveries_weekly_kt"),
    ("Process", "Crop Year to Date Producer Del",                   "process_deliveries_cytd_kt"),
    ("Process", "Stocks at Process Elevators",                      "process_elevator_stocks_kt"),
]
ALBERTA_COL = 3   # 0=label, 1=MB, 2=SK, 3=Alberta, 4=BC, 5=Total

WEEKLY_FLOW_COLS = (
    "producer_deliveries_weekly_kt",
    "primary_shipments_weekly_kt",
    "process_deliveries_weekly_kt",
    "total_deliveries_weekly_kt",
)


# ---------------------------------------------------------------------------
# File reading — uniform interface over xls + xlsx
# ---------------------------------------------------------------------------

def read_sheets_as_rows(path):
    """Return {sheet_name: [(row values, ...), ...]} for either .xls or .xlsx.

    For .xlsx uses openpyxl. For .xls uses xlrd. Both produce the same shape.
    Only reads the Primary and Process sheets (no need for the others).
    """
    suffix = path.suffix.lower()
    target_sheets = {"Primary", "Process"}

    if suffix == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        out = {}
        for name in wb.sheetnames:
            if name in target_sheets:
                out[name] = list(wb[name].iter_rows(values_only=True))
        return out

    if suffix == ".xls":
        wb = xlrd.open_workbook(path)
        out = {}
        for name in wb.sheet_names():
            if name in target_sheets:
                sheet = wb.sheet_by_name(name)
                rows = []
                for r in range(sheet.nrows):
                    row = []
                    for c in range(sheet.ncols):
                        cell = sheet.cell(r, c)
                        # Treat empty cells as None to match openpyxl behaviour
                        if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
                            row.append(None)
                        else:
                            row.append(cell.value)
                    rows.append(tuple(row))
                out[name] = rows
        return out

    raise ValueError(f"Unsupported file type: {suffix}")


# ---------------------------------------------------------------------------
# Sheet-walking helpers (identical to before)
# ---------------------------------------------------------------------------

def find_section_row(rows, needle):
    for i, row in enumerate(rows):
        if row and row[0] and needle in str(row[0]):
            return i
    return None


def find_canola_below(rows, start_idx, max_lookahead=20):
    for i in range(start_idx, min(start_idx + max_lookahead, len(rows))):
        row = rows[i]
        if row and row[0] and str(row[0]).strip() == "Canola":
            return row
    return None


# ---------------------------------------------------------------------------
# Crop-year handling
# ---------------------------------------------------------------------------

def folder_to_crop_year(folder_name):
    """Convert folder name '2016-2017' into crop year string '2016/17'.

    Returns None if the folder name doesn't match the pattern.
    """
    m = re.match(r"(\d{4})-(\d{4})", folder_name)
    if not m:
        return None
    start_year = int(m.group(1))
    end_year = int(m.group(2))
    if end_year != start_year + 1:
        return None
    return f"{start_year}/{end_year % 100:02d}"


def crop_year_for_date(d):
    """Return crop year string for a given date. Aug 1 starts new crop year."""
    if d.month >= 8:
        return f"{d.year}/{(d.year + 1) % 100:02d}"
    return f"{d.year - 1}/{d.year % 100:02d}"


# ---------------------------------------------------------------------------
# Parsing — produces one or two record dicts per file
# ---------------------------------------------------------------------------

def parse_one_file(path, folder_crop_year):
    """Read one GSW file, return list of dicts (1 normally, 2 if combined-week).

    Raises an exception if it can't be parsed — caller logs it.
    """
    sheets = read_sheets_as_rows(path)
    if "Primary" not in sheets:
        raise RuntimeError(f"no 'Primary' sheet found (sheets: {list(sheets.keys())})")

    # Extract the AB canola number from each sub-section
    values = {}
    for sheet_name, needle, col in SECTIONS:
        if sheet_name not in sheets:
            values[col] = None
            continue
        title_idx = find_section_row(sheets[sheet_name], needle)
        if title_idx is None:
            values[col] = None
            continue
        canola_row = find_canola_below(sheets[sheet_name], title_idx)
        if canola_row is None:
            values[col] = None
            continue
        raw = canola_row[ALBERTA_COL] if len(canola_row) > ALBERTA_COL else None
        if raw in (None, "", " "):
            values[col] = None
        else:
            try:
                values[col] = float(raw)
            except (TypeError, ValueError):
                values[col] = None

    # Computed totals (treat NULL as 0 only for the addition — but if BOTH are
    # NULL we want NULL, not 0)
    def _safe_sum(*vs):
        if all(v is None for v in vs):
            return None
        return sum((v or 0) for v in vs)

    values["total_deliveries_weekly_kt"] = _safe_sum(
        values.get("producer_deliveries_weekly_kt"),
        values.get("process_deliveries_weekly_kt"),
    )
    values["total_deliveries_cytd_kt"] = _safe_sum(
        values.get("producer_deliveries_cytd_kt"),
        values.get("process_deliveries_cytd_kt"),
    )
    values["total_visible_stocks_kt"] = _safe_sum(
        values.get("primary_elevator_stocks_kt"),
        values.get("process_elevator_stocks_kt"),
        values.get("condo_storage_kt"),
    )

    # Parse title text to get week number + week_ending date
    title_idx = find_section_row(sheets["Primary"], "Producer Deliveries to Primary Elevators -")
    if title_idx is None:
        raise RuntimeError("can't find 'Producer Deliveries to Primary Elevators -' title row")
    title_text = str(sheets["Primary"][title_idx][0])

    week_match = re.search(r"Week\s+(\d+)", title_text)
    if not week_match:
        raise RuntimeError(f"can't parse week number from title: {title_text!r}")
    week_number = int(week_match.group(1))

    dates = re.findall(r"([A-Z][a-z]+ \d+, \d{4})", title_text)
    if not dates:
        raise RuntimeError(f"can't parse week-ending date from title: {title_text!r}")
    parsed_dates = [datetime.strptime(d, "%B %d, %Y").date() for d in dates]
    week_ending = parsed_dates[-1]

    # Cross-check: does the file's internal date agree with the folder's crop year?
    file_crop_year = crop_year_for_date(week_ending)
    if file_crop_year != folder_crop_year:
        # Use the folder name as the authoritative source per project decision,
        # but log the conflict so we can investigate.
        raise RuntimeError(
            f"folder/file crop year mismatch: folder={folder_crop_year}, "
            f"file title says {week_ending} = {file_crop_year}"
        )

    # Combined-week detection (date range > 10 days = combined file)
    is_combined = False
    span_days = None
    if len(parsed_dates) >= 2:
        span_days = (parsed_dates[-1] - parsed_dates[0]).days
        if span_days >= 10:
            is_combined = True

    if not is_combined:
        return [_build_record(week_number, week_ending, folder_crop_year, values, path.name, notes=None)]

    # Combined-week: split into two rows.
    earlier_week_num = week_number - 1
    earlier_week_end = week_ending - timedelta(days=7)
    note = f"Combined-week file ({span_days}-day span). Source: {path.name}"

    # Later week: stocks + CYTD as published, weekly flows NULL
    later_vals = dict(values)
    for col in WEEKLY_FLOW_COLS:
        later_vals[col] = None
    later_row = _build_record(week_number, week_ending, folder_crop_year, later_vals, path.name,
                              notes=note + " [later half of combined period]")

    # Earlier week: CYTD derived by subtraction, stocks NULL
    earlier_vals = {k: None for k in values.keys()}
    earlier_vals["producer_deliveries_cytd_kt"] = _safe_sub(
        values.get("producer_deliveries_cytd_kt"),
        values.get("producer_deliveries_weekly_kt"),
    )
    earlier_vals["primary_shipments_cytd_kt"] = _safe_sub(
        values.get("primary_shipments_cytd_kt"),
        values.get("primary_shipments_weekly_kt"),
    )
    earlier_vals["process_deliveries_cytd_kt"] = _safe_sub(
        values.get("process_deliveries_cytd_kt"),
        values.get("process_deliveries_weekly_kt"),
    )
    earlier_vals["total_deliveries_cytd_kt"] = _safe_sub(
        values.get("total_deliveries_cytd_kt"),
        values.get("total_deliveries_weekly_kt"),
    )
    earlier_row = _build_record(earlier_week_num, earlier_week_end, folder_crop_year, earlier_vals,
                                path.name,
                                notes=note + " [earlier half - CYTD derived, weekly/stocks NULL]")

    return [earlier_row, later_row]


def _safe_sub(a, b):
    if a is None or b is None:
        return None
    return a - b


def _build_record(week_number, week_ending, crop_year, values, source_file, notes):
    return {
        "week_number": week_number,
        "week_ending": week_ending.isoformat(),
        "crop_year":   crop_year,
        "geography":   "Alberta",
        "crop":        "Canola",
        "source_file": source_file,
        "notes":       notes,
        **{k: values.get(k) for k in (
            "producer_deliveries_weekly_kt", "primary_shipments_weekly_kt",
            "producer_deliveries_cytd_kt", "primary_shipments_cytd_kt",
            "primary_elevator_stocks_kt", "condo_storage_kt",
            "process_deliveries_weekly_kt", "process_deliveries_cytd_kt",
            "process_elevator_stocks_kt",
            "total_deliveries_weekly_kt", "total_deliveries_cytd_kt",
            "total_visible_stocks_kt",
        )},
    }


# ---------------------------------------------------------------------------
# DB plumbing
# ---------------------------------------------------------------------------

def ensure_notes_column(con):
    cur = con.cursor()
    cur.execute("PRAGMA table_info(cgc_weekly)")
    cols = {r[1] for r in cur.fetchall()}
    if "notes" not in cols:
        cur.execute("ALTER TABLE cgc_weekly ADD COLUMN notes TEXT")
        con.commit()
        print("(added 'notes' column to cgc_weekly)")


def insert_record(cur, record):
    cur.execute("""
        INSERT OR REPLACE INTO cgc_weekly (
            week_number, week_ending, crop_year, geography, crop,
            producer_deliveries_weekly_kt, primary_shipments_weekly_kt,
            producer_deliveries_cytd_kt, primary_shipments_cytd_kt,
            primary_elevator_stocks_kt, condo_storage_kt,
            process_deliveries_weekly_kt, process_deliveries_cytd_kt,
            process_elevator_stocks_kt,
            total_deliveries_weekly_kt, total_deliveries_cytd_kt,
            total_visible_stocks_kt,
            source_file, notes
        ) VALUES (
            :week_number, :week_ending, :crop_year, :geography, :crop,
            :producer_deliveries_weekly_kt, :primary_shipments_weekly_kt,
            :producer_deliveries_cytd_kt, :primary_shipments_cytd_kt,
            :primary_elevator_stocks_kt, :condo_storage_kt,
            :process_deliveries_weekly_kt, :process_deliveries_cytd_kt,
            :process_elevator_stocks_kt,
            :total_deliveries_weekly_kt, :total_deliveries_cytd_kt,
            :total_visible_stocks_kt,
            :source_file, :notes
        )
    """, record)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not CGC_DIR.exists():
        print(f"ERROR: directory not found: {CGC_DIR}")
        sys.exit(1)

    # Discover year folders
    year_folders = sorted([p for p in CGC_DIR.iterdir() if p.is_dir()])
    if not year_folders:
        # Maybe flat layout (no subfolders) — try the old behaviour
        flat_files = sorted(list(CGC_DIR.glob("*.xls")) + list(CGC_DIR.glob("*.xlsx")))
        if flat_files:
            print(f"No year folders, but found {len(flat_files)} loose files in {CGC_DIR}")
            print("WARNING: without a year folder I can't tell what crop year these belong to.")
            print("         Move them into a folder named like '2023-2024'.")
            sys.exit(1)
        print(f"No year folders or files found in {CGC_DIR}")
        sys.exit(0)

    # Validate folder names
    valid_folders = []
    for folder in year_folders:
        cy = folder_to_crop_year(folder.name)
        if cy is None:
            print(f"  WARN  skipping folder {folder.name} (not in YYYY-YYYY format)")
            continue
        valid_folders.append((folder, cy))

    print(f"Discovered {len(valid_folders)} year folder(s):")
    for folder, cy in valid_folders:
        n = len(list(folder.glob("*.xls"))) + len(list(folder.glob("*.xlsx")))
        print(f"  {folder.name}  ->  crop year {cy}  ({n} files)")
    print()

    con = sqlite3.connect(DB_PATH)
    ensure_notes_column(con)
    cur = con.cursor()

    # Collect all (path, crop_year) pairs
    all_files = []
    for folder, cy in valid_folders:
        files = sorted(list(folder.glob("*.xls")) + list(folder.glob("*.xlsx")))
        for f in files:
            all_files.append((f, cy))

    print(f"Processing {len(all_files)} files...\n")

    inserted, skipped, failed = 0, 0, 0
    failures = []     # (path, reason)
    progress_every = max(1, len(all_files) // 20)   # ~20 progress lines

    for i, (path, cy) in enumerate(all_files, start=1):
        rel_path = f"{path.parent.name}/{path.name}"
        try:
            records = parse_one_file(path, cy)
            if not records:
                skipped += 1
                failures.append((rel_path, "parse returned no records"))
                continue
            for record in records:
                insert_record(cur, record)
                inserted += 1
            if i % progress_every == 0 or i == len(all_files):
                tag = " [SPLIT]" if len(records) == 2 else ""
                print(f"  [{i:>4}/{len(all_files)}]  {rel_path:<45}  wk {records[-1]['week_number']:>2}  "
                      f"ending {records[-1]['week_ending']}{tag}")
        except Exception as e:
            failed += 1
            failures.append((rel_path, str(e)))
            tb = traceback.format_exc()
            # Show ERR immediately so user sees it without scanning the log
            print(f"  ERR   {rel_path}: {e}")

    con.commit()

    print(f"\n{'='*60}")
    print(f"Done: {inserted} row(s) inserted/replaced from {len(all_files)} file(s)")
    print(f"      {failed} files failed, {skipped} skipped")

    if failures:
        FAIL_LOG.parent.mkdir(parents=True, exist_ok=True)
        with open(FAIL_LOG, "w") as f:
            f.write(f"CGC load failures - {datetime.now().isoformat()}\n")
            f.write(f"{'='*60}\n")
            for path, reason in failures:
                f.write(f"{path}\n  {reason}\n\n")
        print(f"      Failures logged to {FAIL_LOG}")

    # Coverage summary
    print(f"\n--- cgc_weekly coverage by crop year ---")
    cur.execute("""
        SELECT crop_year, COUNT(*) AS weeks,
               MIN(week_number) AS first_wk, MAX(week_number) AS last_wk,
               MIN(week_ending) AS first_end, MAX(week_ending) AS last_end
        FROM cgc_weekly WHERE geography='Alberta' AND crop='Canola'
        GROUP BY crop_year ORDER BY crop_year
    """)
    print(f"{'crop_year':<10} {'weeks':>6}  {'wk_range':>10}  {'date_range'}")
    for r in cur.fetchall():
        print(f"{r[0]:<10} {r[1]:>6}  {r[2]:>3}-{r[3]:<6}  {r[4]} to {r[5]}")

    # Gap detector per crop year
    cur.execute("""
        SELECT crop_year, week_number FROM cgc_weekly
        WHERE geography='Alberta' AND crop='Canola'
        ORDER BY crop_year, week_number
    """)
    by_cy = {}
    for cy, wn in cur.fetchall():
        by_cy.setdefault(cy, []).append(wn)

    any_gaps = False
    for cy, weeks in by_cy.items():
        weeks = sorted(set(weeks))
        if weeks:
            full = set(range(weeks[0], weeks[-1] + 1))
            gaps = sorted(full - set(weeks))
            if gaps:
                if not any_gaps:
                    print()
                any_gaps = True
                # Compact representation if many gaps
                if len(gaps) > 10:
                    print(f"  WARN  {cy} missing {len(gaps)} weeks: {gaps[:5]}...{gaps[-3:]}")
                else:
                    print(f"  WARN  {cy} missing weeks: {gaps}")

    con.close()


if __name__ == "__main__":
    main()
